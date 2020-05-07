# -*- coding: utf-8 -*-
"""
Created on Wed May  6 15:10:54 2020

@author: Carlesso
"""
import glob
import csv
import pandas as pd
from datetime import datetime, timedelta
import pytz
from openpyxl import Workbook



# this file was manually checked
metadata_file = "./stazioni_good.csv"
metadata_df = pd.read_csv(
    filepath_or_buffer = metadata_file, 
    sep=";",
    decimal = ".",
    header =  0,
    usecols = ["code", "regione", "label_good"],
    index_col = 0,
    )

metadata_series = pd.Series(metadata_df["label_good"])

filename_list = glob.glob("input/lmb[0-9][0-9][0-9].csv")
#filename_list = glob.glob("input/lmb080.csv")
output_file = "output/temperatureOrarieMN.xlsx"

one_hour = timedelta(hours=1)
italy = pytz.timezone("Europe/Rome")
utc = pytz.timezone("UTC")

# create an Excel Workbook to hold all stations
# TODO maybe set iso_dates=True 
wb = Workbook(write_only=True)
ws0 = wb.create_sheet("metadata")
ws0.append(["Questo file Excel riporta le temperature di alcune stazioni MeteoNetwork.",])
ws0.append(["italy_dt è l'ora riportata nei dati forniti da MeteoNetwork, che interpretiamo come ora Italiana",])
ws0.append(["solar_dt è invece l'ora UTC+1, quindi indipendente dai periodi in cui vige l'ora legale.",])
for filename in filename_list:
    station_id = filename[6:12]
    station_label = metadata_series[station_id]
    print(station_id, station_label)
    # create a sheet for the station
    ws = wb.create_sheet(title=station_label)
    # append the column header to sheet
    ws.append(["italy_dt", "solar_dt", station_label])
    with open(filename) as file:
        # csv header is 
        # code;datetime;temperature
        reader = csv.DictReader(file, delimiter=";", )
        for row in reader:
            try:
s                italy_dt = italy.localize(naive_dt)
                utc_dt = italy_dt.astimezone(utc)
                solar_dt = utc_dt + one_hour
            except:
                print("not a valid ISO string ", row)
                break
            try:
                temperature = round(float(row["temperature"]), ndigits=1)
            except: 
                temperature = -999.9
            # append a record 
            ws.append( [ italy_dt, solar_dt, temperature ])
wb.save(output_file)
        
        